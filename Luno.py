# Import relevant libraries
import requests
import time
import datetime
from openpyxl import load_workbook
from luno_python.client import Client

#####################################################################
# Make connection to Luno server
#####################################################################

c = Client(api_key_id="cu8mryhtp5efk", api_key_secret="OmvIryVwHWYe9vaEqDLg4kJtXYXVFTLw8nwx-ng-x_Y")
#c = Client(api_key_id="cu8mryhtp5efk", api_key_secret="1yW7o-KuwgkBfmRz_LBJFMeL8lzp070slIVLjrOqVaE")

#####################################################################
# Get tickers
#####################################################################

res = c.get_ticker("XBTZAR")

#####################################################################
# Get balances
# - Commented the below out due to unhandled exceptions
# - We will have to check this out when we begin actual trades
#####################################################################

#balance = c.get_balances()

#####################################################################
# Show balances
#####################################################################

# print(balance)
print(res)       #Debug data

# parse XBTZAR
for key,value in res.items():
    if key == 'timestamp':
       luno_timestamp_epoch = value
       #TO-DO:epoch time conversion for writing to xls
    if key == 'ask':
       luno_askprice = value
       
# Add Indicators to Sheet
wb = load_workbook(filename = 'DATE Check.xlsx')
ws = wb['LunoData']
ws.insert_rows(2)
#ws.cell(2,1,luno_timestamp_epoch)
ws.cell(2,1,time.strftime('%Y-%m-%dT%H:%M:%S %Z',time.localtime(time.time())))
ws.cell(2,2,luno_askprice)
wb.save(filename = 'DATE Check.xlsx')


# try:
#    res = c.get_ticker(pair='XBTZAR')
#    print(res)
# except Exception as e:
#   print (e)