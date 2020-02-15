# Import relevant libraries
import requests
import time
import datetime
import json
from openpyxl import load_workbook

#####################################################################
# Get tickers
#####################################################################

# BTCUSD
res = requests.get("https://www.bitstamp.net/api/ticker/")
res_string = res.content.decode() #convert byte object to str
btcusd = json.loads(res_string) #convert str to dict

# BTCEUR
res = requests.get("https://www.bitstamp.net/api/v2/ticker/btceur")
res_string = res.content.decode() #convert byte object to str
btceur = json.loads(res_string) #convert str to dict

# parse BTCUSD
for key,value in btcusd.items():
    if key == 'timestamp':
       bitstamp_btcusd_timestamp_epoch = value
      #TO-DO:epoch time conversion for writing to xls
    if key == 'bid':
       bitstamp_btcusd_bidprice = value

# parse BTCEUR
for key,value in btceur.items():
    if key == 'timestamp':
       bitstamp_btceur_timestamp_epoch = value
      #TO-DO:epoch time conversion for writing to xls
    if key == 'bid':
       bitstamp_btceur_bidprice = value
       
# Add Indicators to Sheet
wb = load_workbook(filename = 'DATE Check.xlsx')
ws = wb['BitStampData']
ws.insert_rows(2)
#---ws.cell(2,1,bitstamp_btceur_timestamp_epoch)
ws.cell(2,1,time.strftime('%Y-%m-%dT%H:%M:%S %Z',time.localtime(time.time())))
ws.cell(2,2,bitstamp_btceur_bidprice)
ws.cell(2,3,bitstamp_btcusd_bidprice)
wb.save(filename = 'DATE Check.xlsx')
