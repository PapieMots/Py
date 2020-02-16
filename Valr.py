# Import relevant libraries
import requests
import time
import datetime
import json
from openpyxl import load_workbook

#####################################################################
# Get tickers
#####################################################################

# BTCZAR
res = requests.get("https://api.valr.com/v1/public/BTCZAR/marketsummary")
print(res)
print(type(res))

#res_string = res.content.decode() #convert byte object to str
#btczar = json.loads(res_string) #convert str to dict

# parse BTCZAR
#for key,value in btczar.items():
#    if key == 'timestamp':
#       valr_btczar_timestamp_epoch = value
      #TO-DO:epoch time conversion for writing to xls
#    if key == 'ask':
#       valr_btczar_askprice = value


       
# Add Indicators to Sheet
#wb = load_workbook(filename = 'DATE Check.xlsx')
#ws = wb['VALRData']
#ws.insert_rows(2)
#---ws.cell(2,1,bitstamp_btczar_timestamp_epoch)
#ws.cell(2,1,time.strftime('%Y-%m-%dT%H:%M:%S %Z',time.localtime(time.time())))
#ws.cell(2,2,valr_btczar_askprice)
#wb.save(filename = 'DATE Check.xlsx')